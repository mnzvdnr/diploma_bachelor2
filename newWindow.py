import os
import shutil
import sys
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from openpyxl import load_workbook
import os
from datetime import datetime
import shutil

class NewWindow(QMainWindow):
    def __init__(self,f, bool=True, parent=None):
        super().__init__(parent)
        self.mdi = QMdiArea()
        self.setCentralWidget(self.mdi)
        self.setWindowTitle('Просмотр')
        self.file_name = f
        self.bar = self.menuBar()
        if bool:
            file_menu = self.bar.addMenu('Результаты расчёта')
            save_menu = self.bar.addMenu('Сохранить решение')
            action2 = QAction("сохранить", self)
            action2.triggered.connect(self.save)
            save_menu.addAction(action2)
        else:
            file_menu = self.bar.addMenu('Сохранённые решения')
        # Получаем список файлов в папке
        file_list = os.listdir(f)

        # Добавляем каждый файл в меню
        for filename in file_list:
            action = QAction(filename, self)
            action.triggered.connect(lambda _, filename=filename: self.open_file(f, filename))
            file_menu.addAction(action)

    def create_and_move_folder(self):
        # Проверяем, существует ли папка "output"
        if not os.path.exists("output"):
            # Если нет, создаем её
            os.mkdir("output")

        # Переходим в папку "output"
        os.chdir("output")

        # Создаем папку с именем input_folder_name
        os.mkdir(self.file_name)

        # Копируем содержимое папки input_folder_name в output_folder_name
        shutil.copytree(f"../{self.file_name}", self.file_name)
        current_date = datetime.now().strftime('%Y-%m-%d')

        # Создаем имя папки с учетом текущей даты
        output_folder_name = f'{self.file_name}_{current_date}'

        # Переименовываем папку в output_folder_name
        os.rename(self.file_name, output_folder_name)


    def save(self):
        # Проверяем существование папки output
        output_path = 'output'
        if not os.path.exists(output_path):
            os.makedirs(output_path)
        # Получаем текущую дату
        current_date = datetime.now().strftime('%Y-%m-%d')
        # Создаем имя папки с учетом текущей даты
        new_folder_name = f'{self.file_name}_{current_date}'
        # Полный путь к новой папке
        new_folder_path = os.path.join(output_path, new_folder_name)
        # Создаем папку
        if not os.path.exists(new_folder_path):
            os.makedirs(new_folder_path)
        # Копируем содержимое существующей папки в новую папку
        existing_folder_path = self.file_name
        for root, dirs, files in os.walk(existing_folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                shutil.copy(file_path, new_folder_path)
    def open_file(self, folder, filename):
        filepath = os.path.join(folder, filename)
        print("Путь к файлу:", filepath)  # Добавляем эту строку для вывода пути к файлу
        if os.path.isfile(filepath):
            sub = QMdiSubWindow()
            sub.setWindowTitle(filename)
            # Определяем расширение файла
            _, file_extension = os.path.splitext(filename)

            if file_extension == '.png':
                label = QLabel()
                pixmap = QPixmap(filepath)
                label.setPixmap(pixmap)
                sub.setWidget(label)
            elif file_extension == '.xlsx':
                # Открываем файл Excel
                wb = load_workbook(filepath)
                ws = wb.active
                # Получаем данные из ячеек
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)

                # Создаем таблицу для отображения данных
                table = QTableWidget(len(data), len(data[0]), parent=self)
                for i, row in enumerate(data):
                    for j, value in enumerate(row):
                        item = QTableWidgetItem(str(value))
                        table.setItem(i, j, item)
                sub.setWidget(table)

            self.mdi.addSubWindow(sub)
            sub.show()
            # Устанавливаем режим плиток
            self.mdi.tileSubWindows()
        else:
            print("Файл не найден")

